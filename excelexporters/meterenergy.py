import base64
import uuid
import os
from openpyxl.chart import (
    BarChart,
    Reference,
)
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList


####################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excelexporters file
# Step 3: Encode the excelexporters file to Base64
####################################################################################################################

def export(result, name, reporting_start_datetime_local, reporting_end_datetime_local, period_type):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None


    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report, name, reporting_start_datetime_local, reporting_end_datetime_local, period_type):

    wb = Workbook()
    ws = wb.active

    # Row height
    ws.row_dimensions[1].height = 118
    for i in range(2, 11 + 1):
        ws.row_dimensions[i].height = 30

    for i in range(12, 43 + 1):
        ws.row_dimensions[i].height = 30

    # Col width
    ws.column_dimensions['A'].width = 1.5

    for i in range(ord('B'), ord('I')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Constantia', size=15, bold=True)
    title_font = Font(name='宋体', size=15, bold=True)
    data_font = Font(name='Franklin Gothic Book', size=11)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    c_r_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'B1')

    # Title
    ws['B3'].font = name_font
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = 'Name:'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = name

    ws['D3'].font = name_font
    ws['D3'].alignment = b_r_alignment
    ws['D3'] = 'Period:'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'].font = name_font
    ws['E3'] = period_type

    ws['F3'].font = name_font
    ws['F3'].alignment = b_r_alignment
    ws['F3'] = 'Date:'
    ws['G3'].border = b_border
    ws['G3'].alignment = b_c_alignment
    ws['G3'].font = name_font
    ws['G3'] = reporting_start_datetime_local + "__" + reporting_end_datetime_local
    ws.merge_cells("G3:H3")

    if "reporting_period" not in report.keys() or \
            "values" not in report['reporting_period'].keys() or len(report['reporting_period']['values']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename
    #################################################
    # First: 能耗分析
    # 6: title
    # 7: table title
    # 8~9 table_data
    #################################################
    has_energy_data_flag = True

    if "values" not in report['reporting_period'].keys() or len(report['reporting_period']['values']) == 0:
        has_energy_data_flag = False

    if has_energy_data_flag:
        ws['B6'].font = title_font
        ws['B6'] = name + '能耗分析'

        reporting_period_data = report['reporting_period']
        # print(reporting_period_data)
        category = report['meter']['energy_category_name']
        ca_len = len(category)

        ws['B7'].fill = table_fill

        ws['B8'].font = title_font
        ws['B8'].alignment = c_c_alignment
        ws['B8'] = '能耗'
        ws['B8'].border = f_border

        ws['B9'].font = title_font
        ws['B9'].alignment = c_c_alignment
        ws['B9'] = '环比'
        ws['B9'].border = f_border

        for i in range(0, ca_len):
            col = chr(ord('C') + i)
            row = '7'
            cell = col + row
            ws[col + '7'].fill = table_fill
            ws[col + '7'].font = name_font
            ws[col + '7'].alignment = c_c_alignment
            ws[col + '7'] = report['meter']['energy_category_name'] + " (" + report['meter']['unit_of_measure'] + ")"
            ws[col + '7'].border = f_border

            ws[col + '8'].font = name_font
            ws[col + '8'].alignment = c_c_alignment
            ws[col + '8'] = round(reporting_period_data['total_in_category'], 0)
            ws[col + '8'].border = f_border

            ws[col + '9'].font = name_font
            ws[col + '9'].alignment = c_c_alignment
            ws[col + '9'] = str(round(reporting_period_data['increment_rate'] * 100, 2)) + "%" \
                if reporting_period_data['increment_rate'] is not None else "-"
            ws[col + '9'].border = f_border

        # TCE TCO2E
        end_col = col
        # TCE
        tce_col = chr(ord(end_col) + 1)
        ws[tce_col + '7'].fill = table_fill
        ws[tce_col + '7'].font = name_font
        ws[tce_col + '7'].alignment = c_c_alignment
        ws[tce_col + '7'] = "TCE"
        ws[tce_col + '7'].border = f_border

        ws[tce_col + '8'].font = name_font
        ws[tce_col + '8'].alignment = c_c_alignment
        ws[tce_col + '8'] = round(reporting_period_data['total_in_category'], 0)
        ws[tce_col + '8'].border = f_border

        ws[tce_col + '9'].font = name_font
        ws[tce_col + '9'].alignment = c_c_alignment
        ws[tce_col + '9'] = str(round(reporting_period_data['increment_rate'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate'] is not None else "-"
        ws[tce_col + '9'].border = f_border

        # TCO2E
        tco2e_col = chr(ord(end_col) + 2)
        ws[tco2e_col + '7'].fill = table_fill
        ws[tco2e_col + '7'].font = name_font
        ws[tco2e_col + '7'].alignment = c_c_alignment
        ws[tco2e_col + '7'] = "TCO2E"
        ws[tco2e_col + '7'].border = f_border

        ws[tco2e_col + '8'].font = name_font
        ws[tco2e_col + '8'].alignment = c_c_alignment
        ws[tco2e_col + '8'] = round(reporting_period_data['total_in_kgco2e'], 0)
        ws[tco2e_col + '8'].border = f_border

        ws[tco2e_col + '9'].font = name_font
        ws[tco2e_col + '9'].alignment = c_c_alignment
        ws[tco2e_col + '9'] = str(round(reporting_period_data['increment_rate'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate'] is not None else "-"
        ws[tco2e_col + '9'].border = f_border
    else:
        for i in range(6, 9 + 1):
            ws.row_dimensions[i].height = 0.1
    #################################################
    # Second: 能耗详情
    # 11: title
    # 12 ~ 16: chart
    # 18: table title
    # 19~43: table_data
    #################################################
    has_energy_detail_flag = True
    reporting_period_data = report['reporting_period']
    times = reporting_period_data['timestamps']

    if "values" not in report['reporting_period'].keys() or len(report['reporting_period']['values']) == 0:
        has_energy_detail_flag = False

    if has_energy_detail_flag:
        ws['B11'].font = title_font
        ws['B11'] = name + '能耗详情'

        ws['B18'].fill = table_fill
        ws['B18'].font = title_font
        ws['B18'].border = f_border
        ws['B18'].alignment = c_c_alignment
        ws['B18'] = '时间'
        time = times
        has_data = False
        max_row = 0
        if len(time) > 0:
            has_data = True
            max_row = 18 + len(time)

        if has_data:
            for i in range(0, len(time)):
                col = 'B'
                row = str(19 + i)
                # col = chr(ord('B') + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = time[i]
                ws[col + row].border = f_border

            for i in range(0, ca_len):
                # 12 title
                col = chr(ord('C') + i)

                ws[col + '18'].fill = table_fill
                ws[col + '18'].font = title_font
                ws[col + '18'].alignment = c_c_alignment
                ws[col + '18'] = report['meter']['energy_category_name'] + \
                    " (" + report['meter']['unit_of_measure'] + ")"
                ws[col + '18'].border = f_border

                # 13 data
                time = times
                time_len = len(time)

                for j in range(0, time_len):
                    row = str(19 + j)
                    # col = chr(ord('B') + i)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    ws[col + row] = round(reporting_period_data['values'][j], 0)
                    ws[col + row].border = f_border
                # bar
                # 13~: bar
                bar = BarChart()
                labels = Reference(ws, min_col=2, min_row=19, max_row=max_row + 1)
                bar_data = Reference(ws, min_col=3 + i, min_row=18, max_row=max_row + 1)  # openpyxl bug
                bar.add_data(bar_data, titles_from_data=True)
                bar.set_categories(labels)
                bar.height = 5.25  # cm 1.05*5 1.05cm = 30 pt
                bar.width = 18
                # pie.title = "Pies sold by category"
                bar.dLbls = DataLabelList()
                # bar.dLbls.showCatName = True  # 标签显示
                bar.dLbls.showVal = True  # 数量显示
                bar.dLbls.showPercent = True  # 百分比显示
                # s1 = CharacterProperties(sz=1800)     # 图表中字体大小 *100
                chart_col = chr(ord('B') + 2 * i)
                chart_cell = chart_col + str(max_row + 2)
                ws.add_chart(bar, "B12")
    else:
        for i in range(11, 43 + 1):
            ws.row_dimensions[i].height = 0.0
    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
