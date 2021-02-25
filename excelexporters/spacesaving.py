import base64
import uuid
import os
from openpyxl.chart import (
    PieChart,
    LineChart,
    BarChart,
    Reference,
)
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList


####################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file bytes to Base64
####################################################################################################################


def export(report,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None
    print(report)

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type):
    wb = Workbook()
    ws = wb.active

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Constantia', size=15, bold=True)
    title_font = Font(name='宋体', size=15, bold=True)
    data_font = Font(name='Franklin Gothic Book', size=11)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_r_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    img.width = img.width * 0.85
    img.height = img.height * 0.85
    # img = Image("myems.png")
    ws.add_image(img, 'B1')

    # Title
    ws.row_dimensions[3].height = 60

    ws['B3'].font = name_font
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = 'Name:'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = name

    ws['D3'].font = name_font
    ws['D3'].alignment = b_r_alignment
    ws['D3'] = 'Period:'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'].font = name_font
    ws['E3'] = period_type

    ws['F3'].font = name_font
    ws['F3'].alignment = b_r_alignment
    ws['F3'] = 'Date:'
    ws['G3'].alignment = b_c_alignment
    ws['G3'].font = name_font
    ws['G3'] = reporting_start_datetime_local + "__" + reporting_end_datetime_local
    ws.merge_cells("G3:H3")

    if "reporting_period" not in report.keys() or \
            "names" not in report['reporting_period'].keys() or len(report['reporting_period']['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename

    ##################################

    current_row_number = 6

    reporting_period_data = report['reporting_period']

    has_names_data_flag = True

    if "names" not in reporting_period_data.keys() or \
            reporting_period_data['names'] is None or \
            len(reporting_period_data['names']) == 0:
        has_names_data_flag = False

    if has_names_data_flag:
        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' 报告期节约'

        current_row_number += 1

        category = reporting_period_data['names']
        ca_len = len(category)

        ws.row_dimensions[current_row_number].height = 75
        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].border = f_border

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].fill = table_fill
            ws[col + str(current_row_number)].font = name_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = \
                reporting_period_data['names'][i] + " (基线-实际) (" + reporting_period_data['units'][i] + ")"

            col = chr(ord(col) + 1)

        ws[col + str(current_row_number)].fill = table_fill
        ws[col + str(current_row_number)].font = name_font
        ws[col + str(current_row_number)].alignment = c_c_alignment
        ws[col + str(current_row_number)].border = f_border
        ws[col + str(current_row_number)] = '吨标准煤 (基线-实际) (TCE)'

        col = chr(ord(col) + 1)

        ws[col + str(current_row_number)].fill = table_fill
        ws[col + str(current_row_number)].font = name_font
        ws[col + str(current_row_number)].alignment = c_c_alignment
        ws[col + str(current_row_number)].border = f_border
        ws[col + str(current_row_number)] = '吨二氧化碳排放 (基线-实际) (TCO2E)'

        col = chr(ord(col) + 1)

        current_row_number += 1

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = '节约'

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].font = name_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = round(reporting_period_data['subtotals_saving'][i], 2)

            col = chr(ord(col) + 1)

        ws[col + str(current_row_number)].font = name_font
        ws[col + str(current_row_number)].alignment = c_c_alignment
        ws[col + str(current_row_number)].border = f_border
        ws[col + str(current_row_number)] = round(reporting_period_data['total_in_kgce_saving'] / 1000, 2)

        col = chr(ord(col) + 1)

        ws[col + str(current_row_number)].font = name_font
        ws[col + str(current_row_number)].alignment = c_c_alignment
        ws[col + str(current_row_number)].border = f_border
        ws[col + str(current_row_number)] = round(reporting_period_data['total_in_kgco2e_saving'] / 1000, 2)

        col = chr(ord(col) + 1)

        current_row_number += 1

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = '单位面积值'

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].font = name_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = round(reporting_period_data['subtotals_per_unit_area_saving'][i], 2)

            col = chr(ord(col) + 1)

        ws[col + str(current_row_number)].font = name_font
        ws[col + str(current_row_number)].alignment = c_c_alignment
        ws[col + str(current_row_number)].border = f_border
        ws[col + str(current_row_number)] = round(reporting_period_data['total_in_kgce_per_unit_area_saving'] / 1000, 2)

        col = chr(ord(col) + 1)

        ws[col + str(current_row_number)].font = name_font
        ws[col + str(current_row_number)].alignment = c_c_alignment
        ws[col + str(current_row_number)].border = f_border
        ws[col + str(current_row_number)] = \
            round(reporting_period_data['total_in_kgco2e_per_unit_area_saving'] / 1000, 2)

        col = chr(ord(col) + 1)

        current_row_number += 1

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = '环比'

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].font = name_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = str(
                round(reporting_period_data['increment_rates_saving'][i] * 100, 2)) + '%' \
                if reporting_period_data['increment_rates_saving'][i] is not None else '-'

            col = chr(ord(col) + 1)

        ws[col + str(current_row_number)].font = name_font
        ws[col + str(current_row_number)].alignment = c_c_alignment
        ws[col + str(current_row_number)].border = f_border
        ws[col + str(current_row_number)] = str(
            round(reporting_period_data['increment_rate_in_kgce_saving'] * 100, 2)) + '%' \
            if reporting_period_data['increment_rate_in_kgce_saving'] is not None else '-'

        col = chr(ord(col) + 1)

        ws[col + str(current_row_number)].font = name_font
        ws[col + str(current_row_number)].alignment = c_c_alignment
        ws[col + str(current_row_number)].border = f_border
        ws[col + str(current_row_number)] = str(
            round(reporting_period_data['increment_rate_in_kgco2e_saving'] * 100, 2)) + '%' \
            if reporting_period_data['increment_rate_in_kgco2e_saving'] is not None else '-'

        col = chr(ord(col) + 1)

        current_row_number += 2

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' 吨标准煤(TCE)占比'

        current_row_number += 1
        table_start_row_number = current_row_number
        chart_start_row_number = current_row_number

        ws.row_dimensions[current_row_number].height = 60
        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].border = f_border

        ws['C' + str(current_row_number)].fill = table_fill
        ws['C' + str(current_row_number)].font = name_font
        ws['C' + str(current_row_number)].alignment = c_c_alignment
        ws['C' + str(current_row_number)].border = f_border
        ws['C' + str(current_row_number)] = '吨标准煤(TCE)占比'

        current_row_number += 1

        for i in range(0, ca_len):
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = reporting_period_data['names'][i]

            ws['C' + str(current_row_number)].font = name_font
            ws['C' + str(current_row_number)].alignment = c_c_alignment
            ws['C' + str(current_row_number)].border = f_border
            ws['C' + str(current_row_number)] = round(reporting_period_data['subtotals_in_kgce_saving'][i] / 1000, 3)

            current_row_number += 1

        table_end_row_number = current_row_number - 1

        if ca_len < 4:
            current_row_number = current_row_number - ca_len + 4

        current_row_number += 1

        pie = PieChart()
        pie.title = name + ' 吨标准煤(TCE)占比'
        labels = Reference(ws, min_col=2, min_row=table_start_row_number + 1, max_row=table_end_row_number)
        pie_data = Reference(ws, min_col=3, min_row=table_start_row_number, max_row=table_end_row_number)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 7.25
        pie.width = 9
        s1 = pie.series[0]
        s1.dLbls = DataLabelList()
        s1.dLbls.showCatName = False
        s1.dLbls.showVal = True
        s1.dLbls.showPercent = True
        ws.add_chart(pie, 'D' + str(chart_start_row_number))

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' 吨二氧化碳排放(TCO2E)占比'

        current_row_number += 1
        table_start_row_number = current_row_number
        chart_start_row_number = current_row_number

        ws.row_dimensions[current_row_number].height = 60
        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].border = f_border

        ws['C' + str(current_row_number)].fill = table_fill
        ws['C' + str(current_row_number)].font = name_font
        ws['C' + str(current_row_number)].alignment = c_c_alignment
        ws['C' + str(current_row_number)].border = f_border
        ws['C' + str(current_row_number)] = '吨二氧化碳排放(TCO2E)占比'

        current_row_number += 1

        for i in range(0, ca_len):
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = reporting_period_data['names'][i]

            ws['C' + str(current_row_number)].font = name_font
            ws['C' + str(current_row_number)].alignment = c_c_alignment
            ws['C' + str(current_row_number)].border = f_border
            ws['C' + str(current_row_number)] = round(reporting_period_data['subtotals_in_kgco2e_saving'][i] / 1000, 3)

            current_row_number += 1

        table_end_row_number = current_row_number - 1

        if ca_len < 4:
            current_row_number = current_row_number - ca_len + 4

        current_row_number += 1

        pie = PieChart()
        pie.title = name + ' 吨二氧化碳排放(TCO2E)占比'
        labels = Reference(ws, min_col=2, min_row=table_start_row_number + 1, max_row=table_end_row_number)
        pie_data = Reference(ws, min_col=3, min_row=table_start_row_number, max_row=table_end_row_number)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 7.25
        pie.width = 9
        s1 = pie.series[0]
        s1.dLbls = DataLabelList()
        s1.dLbls.showCatName = False
        s1.dLbls.showVal = True
        s1.dLbls.showPercent = True
        ws.add_chart(pie, 'D' + str(chart_start_row_number))

    #############################################

    has_child_space_data_flag = True

    if 'child_space' not in report.keys() or \
            report['child_space'] is None or \
            'energy_category_names' not in report['child_space'].keys() or \
            report['child_space']['energy_category_names'] is None or \
            len(report['child_space']['energy_category_names']) == 0:
        has_child_space_data_flag = False

    if has_child_space_data_flag:
        child_space_data = report['child_space']
        ca_len = len(child_space_data['energy_category_names'])

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' 子空间数据'

        current_row_number += 1
        table_start_row_number = current_row_number

        ws.row_dimensions[current_row_number].height = 60
        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].font = name_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = '子空间'

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].fill = table_fill
            ws[col + str(current_row_number)].font = name_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = \
                child_space_data['energy_category_names'][i] + " (" + child_space_data['units'][i] + ")"
            col = chr(ord(col) + 1)

        current_row_number += 1
        ca_child_len = len(child_space_data['child_space_names_array'][0])

        for i in range(0, ca_child_len):
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = child_space_data['child_space_names_array'][0][i]
            current_row_number += 1

        current_row_number -= ca_child_len

        for i in range(0, ca_child_len):
            col = 'C'
            for j in range(0, ca_len):
                ws[col + str(current_row_number)].font = name_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)].border = f_border
                ws[col + str(current_row_number)] = round(child_space_data['subtotals_saving_array'][j][i], 2)
                col = chr(ord(col) + 1)

            current_row_number += 1

        table_end_row_number = current_row_number - 1

        col = 'B'

        for i in range(0, ca_len):
            pie = PieChart()
            labels = Reference(ws, min_col=2, min_row=table_start_row_number + 1, max_row=table_end_row_number)
            pie_data = Reference(ws, min_col=3 + i, min_row=table_start_row_number, max_row=table_end_row_number)
            pie.add_data(pie_data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = reporting_period_data['names'][i] + " (" + \
                                                            reporting_period_data['units'][i] + ")"
            pie.height = 6.6
            pie.width = 9
            s1 = pie.series[0]
            s1.dLbls = DataLabelList()
            s1.dLbls.showCatName = False
            s1.dLbls.showVal = True
            s1.dLbls.showPercent = True
            chart_cell = ''
            if i % 2 == 0:
                chart_cell = 'B' + str(current_row_number)
            else:
                chart_cell = 'E' + str(current_row_number)
                current_row_number += 5
            ws.add_chart(pie, chart_cell)
            # ws.add_chart(pie, col + str(current_row_number))
            # col = chr(ord(col) + 2)

        if ca_len % 2 == 1:
            current_row_number += 5

        current_row_number += 1

    ################################

    has_values_saving_data = True
    has_timestamps_data = True

    if 'values_saving' not in reporting_period_data.keys() or \
            reporting_period_data['values_saving'] is None or \
            len(reporting_period_data['values_saving']) == 0:
        has_values_saving_data = False

    if 'timestamps' not in reporting_period_data.keys() or \
            reporting_period_data['timestamps'] is None or \
            len(reporting_period_data['timestamps']) == 0 or \
            len(reporting_period_data['timestamps'][0]) == 0:
        has_timestamps_data = False

    if has_values_saving_data and has_timestamps_data:
        ca_len = len(reporting_period_data['names'])
        time = reporting_period_data['timestamps'][0]

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' 详细数据'

        current_row_number += 1

        chart_start_row_number = current_row_number

        current_row_number += ca_len * 6
        table_start_row_number = current_row_number

        ws.row_dimensions[current_row_number].height = 60
        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = '日期时间'

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].fill = table_fill
            ws[col + str(current_row_number)].font = title_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = \
                reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
            col = chr(ord(col) + 1)

        current_row_number += 1

        for i in range(0, len(time)):
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = time[i]

            col = 'C'
            for j in range(0, ca_len):
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)].border = f_border
                ws[col + str(current_row_number)] = round(reporting_period_data['values_saving'][j][i], 2) \
                    if reporting_period_data['values_saving'][j][i] is not None else 0.00
                col = chr(ord(col) + 1)

            current_row_number += 1

        table_end_row_number = current_row_number - 1

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = '小计'

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].font = title_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = round(reporting_period_data['subtotals_saving'][i], 2)
            col = chr(ord(col) + 1)

        current_row_number += 2

        format_time_width_number = 1.0
        min_len_number = 1.0
        min_width_number = 11.0  # format_time_width_number * min_len_number + 4 and min_width_number > 11.0

        if period_type == 'hourly':
            format_time_width_number = 4.0
            min_len_number = 2
            min_width_number = 12.0
        elif period_type == 'daily':
            format_time_width_number = 2.5
            min_len_number = 4
            min_width_number = 14.0
        elif period_type == 'monthly':
            format_time_width_number = 2.1
            min_len_number = 4
            min_width_number = 12.4
        elif period_type == 'yearly':
            format_time_width_number = 1.5
            min_len_number = 5
            min_width_number = 11.5

        for i in range(0, ca_len):
            line = LineChart()
            line.title = '报告期节约 - ' + \
                reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
            labels = Reference(ws, min_col=2, min_row=table_start_row_number + 1, max_row=table_end_row_number)
            line_data = Reference(ws, min_col=3 + i, min_row=table_start_row_number, max_row=table_end_row_number)
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "circle"
            line_data.smooth = True
            line.height = 8.25
            line.width = format_time_width_number * len(time) if len(time) > min_len_number else min_width_number
            if line.width > 24:
                line.width = 24
            line.x_axis.crosses = 'min'
            line.dLbls = DataLabelList()
            line.dLbls.dLblPos = 't'
            line.dLbls.showVal = True
            line.dLbls.showPercent = False
            chart_col = 'B'
            chart_cell = chart_col + str(chart_start_row_number)
            chart_start_row_number += 6
            ws.add_chart(line, chart_cell)

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
