import base64
import uuid
import os
from openpyxl.chart import (
        PieChart,
        BarChart,
        Reference,
    )
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList

####################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file bytes to Base64
####################################################################################################################


def export(result,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    pass
    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type)

    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(result,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type):

    wb = Workbook()
    ws = wb.active

    # Row height
    ws.row_dimensions[1].height = 118
    for i in range(2, 37 + 1):
        ws.row_dimensions[i].height = 30

    for i in range(38, 69 + 1):
        ws.row_dimensions[i].height = 15

    # Col width
    ws.column_dimensions['A'].width = 1.5

    for i in range(ord('B'), ord('I')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Constantia', size=15, bold=True)
    title_font = Font(name='宋体', size=15, bold=True)
    data_font = Font(name='Franklin Gothic Book', size=11)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    c_r_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'B1')

    # Title
    ws['B3'].font = name_font
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = 'Name:'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = result['space']['name']

    ws['D3'].font = name_font
    ws['D3'].alignment = b_r_alignment
    ws['D3'] = 'Period:'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'].font = name_font
    ws['E3'] = period_type

    ws['F3'].font = name_font
    ws['F3'].alignment = b_r_alignment
    ws['F3'] = 'Date:'
    ws['G3'].border = b_border
    ws['G3'].alignment = b_c_alignment
    ws['G3'].font = name_font
    ws['G3'] = reporting_start_datetime_local + "__" + reporting_end_datetime_local
    ws.merge_cells("G3:H3")

    #################################################
    # First: 能耗分析
    # 6: title
    # 7: table title
    # 8~11 table_data
    #################################################

    ws['B6'].font = title_font
    ws['B6'] = '能耗分析'

    report = result['reporting_period']
    print(report)
    category = report['names']
    ca_len = len(category)

    ws['B7'].fill = table_fill

    ws['B8'].font = title_font
    ws['B8'].alignment = c_c_alignment
    ws['B8'] = '能耗'
    ws['B8'].border = f_border

    ws['B9'].font = title_font
    ws['B9'].alignment = c_c_alignment
    ws['B9'] = '单位面积能耗'
    ws['B9'].border = f_border

    ws['B10'].font = title_font
    ws['B10'].alignment = c_c_alignment
    ws['B10'] = '环比'
    ws['B10'].border = f_border

    for i in range(0, ca_len):
        col = chr(ord('C') + i)
        row = '7'
        cell = col + row
        ws[col + '7'].fill = table_fill
        ws[col + '7'].font = name_font
        ws[col + '7'].alignment = c_c_alignment
        ws[col + '7'] = report['names'][i] + " (" + report['units'][i] + ")"
        ws[col + '7'].border = f_border

        ws[col + '8'].font = name_font
        ws[col + '8'].alignment = c_c_alignment
        ws[col + '8'] = round(report['subtotals'][i], 0)
        ws[col + '8'].border = f_border

        ws[col + '9'].font = name_font
        ws[col + '9'].alignment = c_c_alignment
        ws[col + '9'] = round(report['subtotals_per_unit_area'][i], 2)
        ws[col + '9'].border = f_border

        ws[col + '10'].font = name_font
        ws[col + '10'].alignment = c_c_alignment
        ws[col + '10'] = str(round(report['increment_rates'][i] * 100, 2)) + "%"
        ws[col + '10'].border = f_border

    ################################################
    # Second: 分时电耗
    # 12: title
    # 13: table title
    # 14~17 table_data
    ################################################
    ws['B12'].font = title_font
    ws['B12'] = '分时电耗'

    ws['B13'].fill = table_fill
    ws['B13'].font = name_font
    ws['B13'].alignment = c_c_alignment
    ws['B13'].border = f_border

    ws['C13'].fill = table_fill
    ws['C13'].font = name_font
    ws['C13'].alignment = c_c_alignment
    ws['C13'].border = f_border
    ws['C13'] = '分时电耗'

    ws['B14'].font = title_font
    ws['B14'].alignment = c_c_alignment
    ws['B14'] = '尖'
    ws['B14'].border = f_border

    ws['C14'].font = title_font
    ws['C14'].alignment = c_c_alignment
    ws['C14'].border = f_border
    ws['C14'] = round(report['toppeaks'][0], 0)

    ws['B15'].font = title_font
    ws['B15'].alignment = c_c_alignment
    ws['B15'] = '峰'
    ws['B15'].border = f_border

    ws['C15'].font = title_font
    ws['C15'].alignment = c_c_alignment
    ws['C15'].border = f_border
    ws['C15'] = round(report['onpeaks'][0], 0)

    ws['B16'].font = title_font
    ws['B16'].alignment = c_c_alignment
    ws['B16'] = '平'
    ws['B16'].border = f_border

    ws['C16'].font = title_font
    ws['C16'].alignment = c_c_alignment
    ws['C16'].border = f_border
    ws['C16'] = round(report['midpeaks'][0], 0)

    ws['B17'].font = title_font
    ws['B17'].alignment = c_c_alignment
    ws['B17'] = '谷'
    ws['B17'].border = f_border

    ws['C17'].font = title_font
    ws['C17'].alignment = c_c_alignment
    ws['C17'].border = f_border
    ws['C17'] = round(report['offpeaks'][0], 0)

    pie = PieChart()
    labels = Reference(ws, min_col=2, min_row=14, max_row=17)
    pie_data = Reference(ws, min_col=3, min_row=14, max_row=17)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 5.25  # cm 1.05*5 1.05cm = 30 pt
    pie.width = 9
    # pie.title = "Pies sold by category"
    s1 = pie.series[0]
    s1.dLbls = DataLabelList()
    s1.dLbls.showCatName = True  # 标签显示
    s1.dLbls.showVal = True  # 数量显示
    s1.dLbls.showPercent = True  # 百分比显示
    # s1 = CharacterProperties(sz=1800)     # 图表中字体大小 *100

    ws.add_chart(pie, "D13")

    ################################################
    # Third: 子空间能耗
    # 19: title
    # 20: table title
    # 21~24 table_data
    ################################################
    child = result['child_space']
    child_spaces = child['child_space_names_array'][0]
    child_subtotals = child['subtotals_array'][0]

    ws['B19'].font = title_font
    ws['B19'] = '子空间能耗'

    ws['B20'].fill = table_fill
    ws['B20'].border = f_border

    ws['C20'].fill = table_fill
    ws['C20'].font = title_font
    ws['C20'].alignment = c_c_alignment
    ws['C20'].border = f_border
    ws['C20'] = child['energy_category_names'][0]

    ca_len = len(child['energy_category_names'])
    space_len = len(child['child_space_names_array'][0])
    for i in range(0, space_len):
        row = str(i + 21)

        ws['B' + row].font = name_font
        ws['B' + row].alignment = c_c_alignment
        ws['B' + row] = child['child_space_names_array'][0][i]
        ws['B' + row].border = f_border

        for j in range(0, ca_len):
            col = chr(ord('C') + j)
            ws[col + row].font = name_font
            ws[col + row].alignment = c_c_alignment
            ws[col + row] = child['subtotals_array'][0][i]
            ws[col + row].border = f_border
            # pie
            # 25~30: pie
            pie = PieChart()
            labels = Reference(ws, min_col=2, min_row=21, max_row=23)
            pie_data = Reference(ws, min_col=3 + j, min_row=21, max_row=23)
            pie.add_data(pie_data, titles_from_data=True)
            pie.set_categories(labels)
            pie.height = 5.25  # cm 1.05*5 1.05cm = 30 pt
            pie.width = 8
            # pie.title = "Pies sold by category"
            s1 = pie.series[0]
            s1.dLbls = DataLabelList()
            s1.dLbls.showCatName = True  # 标签显示
            s1.dLbls.showVal = True  # 数量显示
            s1.dLbls.showPercent = True  # 百分比显示
            # s1 = CharacterProperties(sz=1800)     # 图表中字体大小 *100
            chart_col = chr(ord('B') + 2 * j)
            chart_cell = chart_col + '26'
            ws.add_chart(pie, chart_cell)

    ################################################
    # Third: 电耗详情
    # 37: title
    # 38: table title
    # 39~69: table_data
    ################################################
    report = result['reporting_period']
    times = report['timestamps']

    ws['B37'].font = title_font
    ws['B37'] = '电耗详情'

    ws['B38'].fill = table_fill
    ws['B38'].border = f_border
    ws['B38'].alignment = c_c_alignment
    ws['B38'] = '时间'
    time = times[0]
    has_data = False
    max_row = 0
    if len(time) > 0:
        has_data = True
        max_row = 38 + len(time)
        print("max_row", max_row)

    if has_data:
        for i in range(0, len(time)):
            col = 'B'
            row = str(39 + i)
            # col = chr(ord('B') + i)
            ws[col + row].font = title_font
            ws[col + row].alignment = c_c_alignment
            ws[col + row] = time[i]
            ws[col + row].border = f_border

        for i in range(0, ca_len):
            # 38 title
            col = chr(ord('C') + i)

            ws[col + '38'].fill = table_fill
            ws[col + '38'].font = title_font
            ws[col + '38'].alignment = c_c_alignment
            ws[col + '38'] = report['names'][i] + " (" + report['units'][i] + ")"
            ws[col + '38'].border = f_border

            # 39 data
            time = times[i]
            time_len = len(time)

            for j in range(0, time_len):
                row = str(39 + j)
                # col = chr(ord('B') + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = round(report['values'][i][j], 0)
                ws[col + row].border = f_border
            # bar
            # 39~: bar
            bar = BarChart()
            labels = Reference(ws, min_col=2, min_row=39, max_row=max_row + 1)
            bar_data = Reference(ws, min_col=3 + i, min_row=38, max_row=max_row + 1)  # openpyxl bug
            bar.add_data(bar_data, titles_from_data=True)
            bar.set_categories(labels)
            bar.height = 5.25  # cm 1.05*5 1.05cm = 30 pt
            bar.width = 18
            # pie.title = "Pies sold by category"
            bar.dLbls = DataLabelList()
            bar.dLbls.showCatName = True  # 标签显示
            bar.dLbls.showVal = True  # 数量显示
            bar.dLbls.showPercent = True  # 百分比显示
            # s1 = CharacterProperties(sz=1800)     # 图表中字体大小 *100
            chart_col = chr(ord('B') + 2 * i)
            chart_cell = chart_col + str(max_row + 2)
            ws.add_chart(bar, chart_cell)

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename

