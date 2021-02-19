import base64
import uuid
import os
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook


####################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file bytes to Base64
####################################################################################################################


def export(report,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None
    print(report)

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type):
    wb = Workbook()
    ws = wb.active

    # Row height
    ws.row_dimensions[1].height = 121

    for i in range(2, 36 + 1):
        ws.row_dimensions[i].height = 30

    for i in range(38, 90 + 1):
        ws.row_dimensions[i].height = 30

    # Col width
    ws.column_dimensions['A'].width = 1.5
    ws.column_dimensions['B'].width = 20.0

    for i in range(ord('C'), ord('I')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Constantia', size=15, bold=True)
    title_font = Font(name='宋体', size=15, bold=True)
    # data_font = Font(name='Franklin Gothic Book', size=11)

    table_fill = PatternFill(fill_type='solid', fgColor='1F496D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    # c_r_alignment = Alignment(vertical='bottom',
    #                           horizontal='center',
    #                           text_rotation=0,
    #                           wrap_text=False,
    #                           shrink_to_fit=False,
    #                           indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    # img = Image("myems.png")
    ws.add_image(img, 'B1')

    # Title
    ws['B3'].font = name_font
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = 'Name:'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = name

    ws['D3'].font = name_font
    ws['D3'].alignment = b_r_alignment
    ws['D3'] = 'Period:'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'].font = name_font
    ws['E3'] = period_type

    ws['F3'].font = name_font
    ws['F3'].alignment = b_r_alignment
    ws['F3'] = 'Date:'
    ws['G3'].border = b_border
    ws['G3'].alignment = b_c_alignment
    ws['G3'].font = name_font
    ws['G3'] = reporting_start_datetime_local + "__" + reporting_end_datetime_local
    ws.merge_cells("G3:H3")

    if "reporting_period" not in report.keys() or \
            "names" not in report['reporting_period'].keys() or len(report['reporting_period']['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename
    #################################################
    # First: 统计分析
    # 5: title
    # 6: table title
    # 6~12 table_data
    # Total: 8 rows
    # if has not energy data: set low height for rows
    #################################################
    reporting_period_data = report['reporting_period']

    has_energy_data_flag = True

    if "names" not in reporting_period_data.keys() or \
            reporting_period_data['names'] is None or \
            len(reporting_period_data['names']) == 0:
        has_energy_data_flag = False

        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename

    if has_energy_data_flag:
        ws['B5'].font = title_font
        ws['B5'] = name + ' 统计分析'

        category = reporting_period_data['names']

        # table_title
        ws['B6'].fill = table_fill
        ws['B6'].font = title_font
        ws['B6'].alignment = c_c_alignment
        ws['B6'] = '报告期'
        ws['B6'].border = f_border

        ws['C6'].font = title_font
        ws['C6'].alignment = c_c_alignment
        ws['C6'] = '算术平均数'
        ws['C6'].border = f_border

        ws['D6'].font = title_font
        ws['D6'].alignment = c_c_alignment
        ws['D6'] = '中位数'
        ws['D6'].border = f_border

        ws['E6'].font = title_font
        ws['E6'].alignment = c_c_alignment
        ws['E6'] = '最小值'
        ws['E6'].border = f_border

        ws['F6'].font = title_font
        ws['F6'].alignment = c_c_alignment
        ws['F6'] = '最大值'
        ws['F6'].border = f_border

        ws['G6'].font = title_font
        ws['G6'].alignment = c_c_alignment
        ws['G6'] = '样本标准差'
        ws['G6'].border = f_border

        ws['H6'].font = title_font
        ws['H6'].alignment = c_c_alignment
        ws['H6'] = '样本方差'
        ws['H6'].border = f_border

        # table_data
        for i, value in enumerate(category):
            row = i * 2 + 7
            ws['B' + str(row)].font = name_font
            ws['B' + str(row)].alignment = c_c_alignment
            ws['B' + str(row)] = reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + " )"
            ws['B' + str(row)].border = f_border

            ws['B' + str(row + 1)].font = name_font
            ws['B' + str(row + 1)].alignment = c_c_alignment
            ws['B' + str(row + 1)] = "环比"
            ws['B' + str(row + 1)].border = f_border

            ws['C' + str(row)].font = name_font
            ws['C' + str(row)].alignment = c_c_alignment
            ws['C' + str(row)] = round(reporting_period_data['means'][i], 2) \
                if reporting_period_data['means'][i] is not None else ''
            ws['C' + str(row)].border = f_border

            ws['C' + str(row + 1)].font = name_font
            ws['C' + str(row + 1)].alignment = c_c_alignment
            ws['C' + str(row + 1)] = str(round(reporting_period_data['means_increment_rate'][i] * 100, 2)) + "%" \
                if reporting_period_data['means_increment_rate'][i] is not None else '0.00%'
            ws['C' + str(row + 1)].border = f_border

            ws['D' + str(row)].font = name_font
            ws['D' + str(row)].alignment = c_c_alignment
            ws['D' + str(row)] = round(reporting_period_data['medians'][i], 2) \
                if reporting_period_data['medians'][i] is not None else ''
            ws['D' + str(row)].border = f_border

            ws['D' + str(row + 1)].font = name_font
            ws['D' + str(row + 1)].alignment = c_c_alignment
            ws['D' + str(row + 1)] = str(round(reporting_period_data['medians_increment_rate'][i] * 100, 2)) + "%" \
                if reporting_period_data['medians_increment_rate'][i] is not None else '0.00%'
            ws['D' + str(row + 1)].border = f_border

            ws['E' + str(row)].font = name_font
            ws['E' + str(row)].alignment = c_c_alignment
            ws['E' + str(row)] = round(reporting_period_data['minimums'][i], 2) \
                if reporting_period_data['minimums'][i] is not None else ''
            ws['E' + str(row)].border = f_border

            ws['E' + str(row + 1)].font = name_font
            ws['E' + str(row + 1)].alignment = c_c_alignment
            ws['E' + str(row + 1)] = str(round(reporting_period_data['minimums_increment_rate'][i] * 100, 2)) + "%" \
                if reporting_period_data['minimums_increment_rate'][i] is not None else '0.00%'
            ws['E' + str(row + 1)].border = f_border

            ws['F' + str(row)].font = name_font
            ws['F' + str(row)].alignment = c_c_alignment
            ws['F' + str(row)] = round(reporting_period_data['maximums'][i], 2) \
                if reporting_period_data['maximums'][i] is not None else ''
            ws['F' + str(row)].border = f_border

            ws['F' + str(row + 1)].font = name_font
            ws['F' + str(row + 1)].alignment = c_c_alignment
            ws['F' + str(row + 1)] = str(round(reporting_period_data['maximums_increment_rate'][i] * 100, 2)) + "%" \
                if reporting_period_data['maximums_increment_rate'][i] is not None else '0.00%'
            ws['F' + str(row + 1)].border = f_border

            ws['G' + str(row)].font = name_font
            ws['G' + str(row)].alignment = c_c_alignment
            ws['G' + str(row)] = round(reporting_period_data['stdevs'][i], 2) \
                if reporting_period_data['stdevs'][i] is not None else ''
            ws['G' + str(row)].border = f_border

            ws['G' + str(row + 1)].font = name_font
            ws['G' + str(row + 1)].alignment = c_c_alignment
            ws['G' + str(row + 1)] = str(round(reporting_period_data['stdevs_increment_rate'][i] * 100, 2)) + "%" \
                if reporting_period_data['stdevs_increment_rate'][i] is not None else '0.00%'
            ws['G' + str(row + 1)].border = f_border

            ws['H' + str(row)].font = name_font
            ws['H' + str(row)].alignment = c_c_alignment
            ws['H' + str(row)] = round(reporting_period_data['variances'][i], 2) \
                if reporting_period_data['variances'][i] is not None else ''
            ws['H' + str(row)].border = f_border

            ws['H' + str(row + 1)].font = name_font
            ws['H' + str(row + 1)].alignment = c_c_alignment
            ws['H' + str(row + 1)] = str(round(reporting_period_data['variances_increment_rate'][i] * 100, 2)) + "%" \
                if reporting_period_data['variances_increment_rate'][i] is not None else '0.00%'
            ws['H' + str(row + 1)].border = f_border

    #################################################
    # Second: 报告期消耗 
    # 14~14+row_title: chart
    # 15+row_title: table title
    # 15+row_title~: table_data
    # Total: 1+row_title+time_len rows
    ################################################

    has_timestamps_flag = True
    if "timestamps" not in reporting_period_data.keys() or \
            reporting_period_data['timestamps'] is None or \
            len(reporting_period_data['timestamps']) == 0:
        has_timestamps_flag = False

    if has_timestamps_flag:
        timestamps = reporting_period_data['timestamps'][0]
        values = reporting_period_data['values']
        names = reporting_period_data['names']
        ca_len = len(names)
        time_len = len(timestamps)
        # title
        row_title = 10 * ca_len
        ws['B' + str(14+row_title)].font = title_font
        ws['B' + str(14+row_title)] = name + ' 详细数据'
        # table_title
        ws['B' + str(15+row_title)].fill = table_fill
        ws['B' + str(15+row_title)].font = name_font
        ws['B' + str(15+row_title)].alignment = c_c_alignment
        ws['B' + str(15+row_title)] = "时间"
        ws['B' + str(15+row_title)].border = f_border

        for i in range(0, ca_len):
            col = chr(ord('C') + i)

            ws[col + str(15+row_title)].font = name_font
            ws[col + str(15+row_title)].alignment = c_c_alignment
            ws[col + str(15+row_title)] = names[i] + " (" + reporting_period_data['units'][i] + ")"
            ws[col + str(15+row_title)].border = f_border
        # table_date
        for i in range(0, time_len):
            rows = i + 16 + 10 * ca_len

            ws['B' + str(rows)].font = name_font
            ws['B' + str(rows)].alignment = c_c_alignment
            ws['B' + str(rows)] = timestamps[i]
            ws['B' + str(rows)].border = f_border

            for index in range(0, ca_len):
                col = chr(ord('C') + index)

                ws[col + str(rows)].font = name_font
                ws[col + str(rows)].alignment = c_c_alignment
                ws[col + str(rows)] = values[index][i]
                ws[col + str(rows)].border = f_border

        # 小计
        row_subtotals = 16 + time_len + 10 * ca_len
        ws['B' + str(row_subtotals)].font = name_font
        ws['B' + str(row_subtotals)].alignment = c_c_alignment
        ws['B' + str(row_subtotals)] = "小计"
        ws['B' + str(row_subtotals)].border = f_border

        for i in range(0, ca_len):
            col = chr(ord('C') + i)

            ws[col + str(row_subtotals)].font = name_font
            ws[col + str(row_subtotals)].alignment = c_c_alignment
            ws[col + str(row_subtotals)] = reporting_period_data['subtotals'][i]
            ws[col + str(row_subtotals)].border = f_border

        # LineChart
        for i in range(0, ca_len):

            lc = LineChart()
            lc.title = "报告期消耗" + names[i] + " (" + reporting_period_data['units'][i] + ")"
            lc.style = 10
            lc.height = 8.40  # cm 1.05*8 1.05cm = 30 pt
            lc.width = 31
            lc.x_axis.majorTickMark = 'in'
            lc.y_axis.majorTickMark = 'in'
            times = Reference(ws, min_col=2, min_row=16+row_title, max_row=15 + row_title + time_len)
            lc_data = Reference(ws, min_col=3 + i, min_row=15+row_title, max_row=15 + row_title + time_len)
            lc.add_data(lc_data, titles_from_data=True)
            lc.set_categories(times)
            ser = lc.series[0]
            ser.marker.symbol = "diamond"
            ser.marker.size = 5
            ws.add_chart(lc, 'B' + str(14 + 10 * i))

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
