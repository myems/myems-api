import base64
import uuid
import os
from openpyxl.chart import (
    BarChart,
    Reference,
)
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList


####################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excelexporters file
# Step 3: Encode the excelexporters file to Base64
####################################################################################################################

def export(result, space_name):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              space_name)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report, space_name):

    wb = Workbook()
    ws = wb.active

    # Row height
    ws.row_dimensions[1].height = 118
    for i in range(2, 5000 + 1):
        ws.row_dimensions[i].height = 30
    # Col width
    ws.column_dimensions['A'].width = 1

    # Font
    name_font = Font(name='Constantia', size=15, bold=True)
    title_font = Font(name='宋体', size=15, bold=True)
    data_font = Font(name='Franklin Gothic Book', size=11)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    c_r_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    for i in range(ord('B'), ord('G')):
        ws.column_dimensions[chr(i)].width = 40.0

    # Img
    ws.merge_cells("B1:F1")
    ws.merge_cells("B2:F2")
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'B1')

    # Title
    ws['B3'].border = f_border
    ws['B3'].font = name_font
    ws['B3'].alignment = b_c_alignment
    ws['B3'] = '名称'

    ws['C3'].border = f_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = '空间'

    ws['D3'].border = f_border
    ws['D3'].font = name_font
    ws['D3'].alignment = b_c_alignment
    ws['D3'] = '成本中心:'

    ws['E3'].border = f_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'].font = name_font
    ws['E3'] = '能耗分类'

    ws['F3'].border = f_border
    ws['F3'].font = name_font
    ws['F3'].alignment = b_c_alignment
    ws['F3'] = ' 描述'

    #Converted to excel
    def pa(i,current_row_number):
        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)] = report['meters'][i]['meter_name']

        ws['C' + str(current_row_number)].font = title_font
        ws['C' + str(current_row_number)].border = f_border
        ws['C' + str(current_row_number)].alignment = c_c_alignment
        ws['C' + str(current_row_number)] = report['meters'][i]['space_name']

        ws['D' + str(current_row_number)].font = title_font
        ws['D' + str(current_row_number)].border = f_border
        ws['D' + str(current_row_number)].alignment = c_c_alignment
        ws['D' + str(current_row_number)] = report['meters'][i]['cost_center_name']

        ws['E' + str(current_row_number)].font = title_font
        ws['E' + str(current_row_number)].border = f_border
        ws['E' + str(current_row_number)].alignment = c_c_alignment
        ws['E' + str(current_row_number)] = report['meters'][i]['energy_category_name']

        ws['F' + str(current_row_number)].font = title_font
        ws['F' + str(current_row_number)].border = f_border
        ws['F' + str(current_row_number)].alignment = c_c_alignment
        ws['F' + str(current_row_number)] = report['meters'][i]['description']


    current_row_number = 4
    for i in range(0,len(report['meters'])):
        if space_name == "远洋太古里":
            if report['meters'][i]['space_name'] == "租区"or"主力店" or "燃气餐饮" or "零售" or "电餐饮" or "高照度" or "其他业态"\
                    or "公区商场" or "给排水" or "扶梯直梯" or "照明及插座" or "空调水" or "空调风" or "特殊功能房间" or "其他用电设备"\
                    or "潜水泵、消防泵等" or "厨房一体化污水设备" or "无障碍升降平台" or "直梯" or "扶梯" or "商场普通照明" or "商场应急照明" or "室外景观照明" or "活动临时电"\
                    or "制冷站" or "锅炉房"or "新风机" or "排风机" or "风机盘管" or "卫生间" or "消防控制室" or "雨水处理机房" or "电视屏幕墙控制机房" \
                    or "地下室弱电进线室" or "商业生活水泵房" or "水景预留机房" or "数据中心" or "电热水器" or "厨房保障电源":
                pa(i, current_row_number)

        elif space_name == "租区":
            if report['meters'][i]['space_name']=="主力店" or "燃气餐饮" or "零售" or "电餐饮"or "高照度" or "其他业态":
               pa(i,current_row_number)

        elif space_name == "公区商场":
            if report['meters'][i]['space_name'] == "给排水" or "扶梯直梯" or "照明及插座" or "空调水" or "空调风" or "特殊功能房间" or "其他用电设备" \
                    or "潜水泵、消防泵等" or "厨房一体化污水设备" \
                    or "无障碍升降平台" or "直梯" or "扶梯" \
                    or "商场普通照明" or "商场应急照明" or "室外景观照明" or "活动临时电" \
                    or "制冷站" or "锅炉房" \
                    or "新风机" or "排风机" or "风机盘管" \
                    or "卫生间" or "消防控制室" or "雨水处理机房" or "电视屏幕墙控制机房" or "地下室弱电进线室" or "商业生活水泵房" or "水景预留机房" or "数据中心" \
                    or "电热水器" or "厨房保障电源" or "公区车库" or "车库通风" or "车库照明" or "应急照明" or "普通照明":
                pa(i,current_row_number)

        elif space_name == "给排水":
            if report['meters'][i]['space_name'] =="潜水泵、消防泵等" or "厨房一体化污水设备":
                pa(i,current_row_number)

        elif space_name == "扶梯直梯":
            if report['meters'][i]['space_name'] =="无障碍升降平台" or "直梯" or "扶梯" :
                pa(i,current_row_number)

        elif space_name == "照明及插座":
            if report['meters'][i]['space_name'] =="商场普通照明" or "商场应急照明" or "室外景观照明" or "活动临时电":
                pa(i,current_row_number)

        elif space_name == "空调水":
            if report['meters'][i]['space_name'] =="制冷站" or "锅炉房":
                pa(i,current_row_number)

        elif space_name == "空调风":
            if report['meters'][i]['space_name'] =="新风机" or "排风机" or "风机盘管":
                pa(i,current_row_number)

        elif space_name == "特殊功能房间":
            if report['meters'][i]['space_name'] =="卫生间" or "消防控制室" or "雨水处理机房" or "电视屏幕墙控制机房" or "地下室弱电进线室" or "商业生活水泵房" or "水景预留机房" or "数据中心":
                pa(i,current_row_number)

        elif space_name == "其他用电设备":
            if report['meters'][i]['space_name'] =="电热水器" or "厨房保障电源":
                pa(i,current_row_number)

        elif space_name == "公区车库":
            if report['meters'][i]['space_name'] =="车库通风" or "车库照明" or "应急照明" or "普通照明":
                pa(i,current_row_number)

        elif space_name == "车库照明":
            if report['meters'][i]['space_name'] == "应急照明" or "普通照明":
                pa(i,current_row_number)

        if space_name == report['meters'][i]['space_name']:
            pa(i,current_row_number)

        current_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename



